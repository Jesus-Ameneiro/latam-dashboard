import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, date, timedelta
import re, copy, base64, io
from openpyxl import load_workbook
import requests

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Ruvixx · Case Investigation", page_icon="🔶",
                   layout="wide", initial_sidebar_state="collapsed")

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
ORG, GRN, RED = "#F97316", "#16A34A", "#EF4444"

COUNTRY_FIX = {
    "domican republic":   "Dominican Republic",
    "dominican repbulic": "Dominican Republic",
    "belice": "Belize", "bolivar": "Bolivia", "ecuardor": "Ecuador",
}
COUNTRY_PILLS = {
    "Mexico":"MX","Colombia":"CO","Ecuador":"EC","Guatemala":"GT",
    "Dominican Republic":"DO","Costa Rica":"CR","El Salvador":"SV",
    "Panama":"PA","Honduras":"HN","Nicaragua":"NI","Belize":"BZ",
    "Argentina":"AR","Chile":"CL","Brazil":"BR","Uruguay":"UY",
    "Paraguay":"PY","Bolivia":"BO","Peru":"PE","Venezuela":"VE",
}
ALL_KNOWN_COUNTRIES = sorted(COUNTRY_PILLS.keys())
MONTH_MAP = {m: i+1 for i, m in enumerate([
    "january","february","march","april","may","june",
    "july","august","september","october","november","december",
])}
DISQ_RE = re.compile(
    r"\b(disqualif|disqualified|reject|rjected|duplicad[ao]?|repeated|"
    r"duplicate\s+of|duplicado|related|case\s+related|already\s+contacted|"
    r"entity\s+already|caso\s+relacionado|caso\s+duplicado|not\s+valid|"
    r"no\s+aplica|kasznar|kaznar|attended\s+by\s+kaz|"
    r"same\s+entity|rjected|repeated\s+case|already\s+processed)\b",
    re.IGNORECASE,
)

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT REGION CONFIG
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_REGIONS = {
    "MCC": {
        "name": "México Central Caribe", "contact": "Tatiana Romero",
        "groups": [
            {"label": "Mexico",                   "countries": ["Mexico"],                                   "quota": 40},
            {"label": "DR, Panama, Costa Rica",   "countries": ["Costa Rica","Dominican Republic","Panama"], "quota": 40},
            {"label": "Nicaragua, GT, SV, BZ, HN","countries": ["Nicaragua","Guatemala","El Salvador","Belize","Honduras"], "quota": 40},
        ],
        # daily_ideal=8 (goal), daily_min=5 (critical threshold ≤5)
        # weekly_ideal=40 (goal), weekly_min=25 (critical threshold ≤25)
        "daily_min": 5, "daily_ideal": 8, "weekly_min": 25, "weekly_ideal": 40, "support": ["Luis"],
    },
    "CS": {
        "name": "Cono Sur", "contact": "Ignacio Duce",
        "groups": [
            {"label": "Argentina",          "countries": ["Argentina"],         "quota": 25},
            {"label": "Chile",              "countries": ["Chile"],             "quota": 20},
            {"label": "Ecuador, Colombia",  "countries": ["Ecuador","Colombia"],"quota": 35},
            {"label": "Peru",               "countries": ["Peru"],              "quota": 10},
            {"label": "Uruguay, Bolivia, Paraguay","countries": ["Uruguay","Bolivia","Paraguay"],"quota": 20},
        ],
        "daily_min": 5, "daily_ideal": 8, "weekly_min": 25, "weekly_ideal": 40, "support": [],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
EMPTY_DF = pd.DataFrame(columns=["date","case_id","country","investigator",
                                  "week_num","region","month_key"])
for k, v in [
    ("all_cases",        []),
    ("all_weeks",        []),
    ("current_week_idx", 0),
    ("prev_week_idx",    0),
    ("view",             "current"),   # "prev" | "current" | "full_month"
    ("full_month_key",   None),
    ("tab",              "MCC"),
    ("dark",             True),
    ("rcfg",             copy.deepcopy(DEFAULT_REGIONS)),
    ("_prev_dark",       None),
    # ── xlsx fetch state ──────────────────────────────────────────────────────
    ("xlsx_fetch_ok",    False),       # True once xlsx loaded from GitHub
    ("xlsx_last_sha",    None),        # GitHub SHA of the last fetched xlsx
    ("xlsx_fetch_ts",    None),        # ISO timestamp of last successful fetch
    ("xlsx_fetch_err",   None),        # last error message, or None
    # ── Batch history (from prebatch-delivery GitHub repo) ────────────────────
    ("batch_history",    None),
    ("delivered_ids",    {"MCC": set(), "CS": set()}),
    ("batch_fetch_ok",   False),
]:
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# THEME
# ─────────────────────────────────────────────────────────────────────────────
dark = st.session_state.dark
st.session_state._prev_dark = dark

BG   = "#1A1614" if dark else "#FEF9F5"
CARD = "#242120" if dark else "#FFFFFF"
BORD = "#3D3532" if dark else "#FED7AA"
TX   = "#FAFAF9" if dark else "#1C1917"
TX2  = "#A8A29E" if dark else "#78716C"
OL   = "#431407" if dark else "#FEF3EA"
OB   = "#7C2D12" if dark else "#FED7AA"
PLT  = "plotly_dark" if dark else "plotly_white"
ABSC = "#44403C" if dark else "#FEE2CC"

# ─────────────────────────────────────────────────────────────────────────────
# BATCH HISTORY — fetch from GitHub prebatch-delivery repo
# Repo:  Jesus-Ameneiro/prebatch-delivery
# File:  batch_history.json
# Struct: {"MCC":[{batch_number,delivery_date,region,profile,total_cases,
#                  cases:[[case_id,name],...]},...], "CS":[...]}
# Case IDs may be compound: "4514228#1,4475779#1" → split on comma.
# ─────────────────────────────────────────────────────────────────────────────
_BATCH_REPO = "Jesus-Ameneiro/prebatch-delivery"
_BATCH_PATH = "batch_history.json"

def fetch_batch_history():
    """Fetch batch_history.json from GitHub and populate delivered_ids sets."""
    token = st.secrets.get("GITHUB_TOKEN", "")
    repo  = st.secrets.get("GITHUB_BATCH_REPO", _BATCH_REPO)
    path  = st.secrets.get("GITHUB_BATCH_PATH", _BATCH_PATH)

    if not token:
        st.session_state.batch_fetch_ok = False
        return False

    url     = f"https://api.github.com/repos/{repo}/contents/{path}"
    headers = {
        "Authorization": f"token {token}",
        "Accept":        "application/vnd.github.v3.raw",
        "User-Agent":    "Ruvixx-Dashboard",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=20)
        if resp.status_code != 200:
            st.session_state.batch_fetch_ok = False
            return False
        data = resp.json()
    except Exception:
        st.session_state.batch_fetch_ok = False
        return False

    st.session_state.batch_history  = data
    st.session_state.delivered_ids  = _build_delivered_set(data)
    st.session_state.batch_fetch_ok = True
    return True


def _build_delivered_set(batch_data):
    """
    Build {region: set_of_case_ids} from the batch history JSON.
    Handles compound case IDs (comma-separated) within a single entry.
    """
    delivered = {"MCC": set(), "CS": set()}
    for region in ("MCC", "CS"):
        for batch in batch_data.get(region, []):
            for entry in batch.get("cases", []):
                raw_id = str(entry[0] if entry else "").strip()
                # Split compound IDs: "4514228#1,4475779#1" → ["4514228#1","4475779#1"]
                for cid in re.split(r",\s*", raw_id):
                    cid = cid.strip()
                    if cid:
                        delivered[region].add(cid)
    return delivered


def batch_history_summary(region):
    """
    Return a sorted list of batch dicts for display (newest first),
    plus a total delivered count for the region.
    """
    bh = st.session_state.batch_history
    if not bh:
        return [], 0
    batches = sorted(
        bh.get(region, []),
        key=lambda b: b.get("batch_number", 0),
        reverse=True,
    )
    total_delivered = sum(b.get("total_cases", 0) for b in batches)
    return batches, total_delivered


# ─────────────────────────────────────────────────────────────────────────────
# XLSX FETCH — download from GitHub latam-dashboard repo
# Repo:  Jesus-Ameneiro/latam-dashboard
# File:  data/LATAM_Internal_Case_Revision.xlsx
# The Apps Script pushes a fresh copy every hour (03:00–23:59 Bogotá).
# The dashboard fetches on startup and on manual Refresh.
# ─────────────────────────────────────────────────────────────────────────────
_XLSX_REPO = "Jesus-Ameneiro/latam-dashboard"
_XLSX_PATH = "data/LATAM_Internal_Case_Revision.xlsx"

def fetch_xlsx_from_github(force=False):
    """
    Download the xlsx from GitHub and call load_xlsx().
    - On first load (xlsx_fetch_ok=False): always fetch.
    - On Refresh (force=True): always fetch.
    - Checks SHA: skips re-parse if file hasn't changed since last fetch.
    Returns True on success, False on any error.
    """
    token = st.secrets.get("GITHUB_TOKEN", "")
    repo  = st.secrets.get("GITHUB_XLSX_REPO", _XLSX_REPO)
    path  = st.secrets.get("GITHUB_XLSX_PATH", _XLSX_PATH)

    if not token:
        st.session_state.xlsx_fetch_err = "GITHUB_TOKEN not set in Streamlit secrets."
        return False

    # ── Step 1: get file metadata (SHA + download_url) ────────────────────────
    meta_url = f"https://api.github.com/repos/{repo}/contents/{path}"
    headers  = {
        "Authorization": f"token {token}",
        "Accept":        "application/vnd.github.v3+json",
        "User-Agent":    "Ruvixx-Dashboard",
    }
    try:
        meta_resp = requests.get(meta_url, headers=headers, timeout=20)
        if meta_resp.status_code == 404:
            st.session_state.xlsx_fetch_err = (
                f"File not found: {repo}/{path}. "
                "Run setupHourlyTrigger() in Apps Script to start pushing the file.")
            return False
        if meta_resp.status_code != 200:
            st.session_state.xlsx_fetch_err = (
                f"GitHub API error {meta_resp.status_code} fetching xlsx metadata.")
            return False
        meta = meta_resp.json()
    except Exception as e:
        st.session_state.xlsx_fetch_err = f"Network error fetching xlsx: {e}"
        return False

    current_sha  = meta.get("sha", "")
    download_url = meta.get("download_url", "")

    # ── Step 2: skip download if SHA unchanged (unless force=True) ────────────
    if (not force
            and st.session_state.xlsx_fetch_ok
            and current_sha == st.session_state.xlsx_last_sha):
        # File unchanged — no need to re-parse
        return True

    # ── Step 3: download raw binary via download_url ──────────────────────────
    if not download_url:
        st.session_state.xlsx_fetch_err = "No download_url in GitHub API response."
        return False

    try:
        dl_resp = requests.get(
            download_url,
            headers={"Authorization": f"token {token}", "User-Agent": "Ruvixx-Dashboard"},
            timeout=60,
        )
        if dl_resp.status_code != 200:
            st.session_state.xlsx_fetch_err = (
                f"Failed to download xlsx: HTTP {dl_resp.status_code}")
            return False
        xlsx_bytes = dl_resp.content
    except Exception as e:
        st.session_state.xlsx_fetch_err = f"Network error downloading xlsx: {e}"
        return False

    # ── Step 4: parse with load_xlsx ──────────────────────────────────────────
    file_obj = io.BytesIO(xlsx_bytes)
    ok = load_xlsx(file_obj)
    if ok:
        st.session_state.xlsx_fetch_ok  = True
        st.session_state.xlsx_last_sha  = current_sha
        st.session_state.xlsx_fetch_ts  = datetime.now().strftime("%b %d, %H:%M")
        st.session_state.xlsx_fetch_err = None
    return ok


# ─────────────────────────────────────────────────────────────────────────────
# REGION HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def get_all_assigned():
    return {c: (rk, gi)
            for rk, rc in st.session_state.rcfg.items()
            for gi, g  in enumerate(rc["groups"])
            for c in g["countries"]}

def total_quota(rk):
    return sum(g["quota"] for g in st.session_state.rcfg[rk]["groups"])

def region_pills(rk):
    return [COUNTRY_PILLS.get(c, c[:2].upper())
            for g in st.session_state.rcfg[rk]["groups"] for c in g["countries"]]

# ─────────────────────────────────────────────────────────────────────────────
# TEXT / DATE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def norm_country(c):
    if not c or str(c).strip() in ("","nan","None"): return ""
    t = str(c).strip()
    return COUNTRY_FIX.get(t.lower(), t)

def is_disq(qa):
    return bool(DISQ_RE.search(str(qa or "")))

def parse_date_val(val):
    if val is None: return None
    try:
        if isinstance(val, pd.Timestamp):
            return val.strftime("%Y-%m-%d") if pd.notna(val) else None
        if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
        if isinstance(val, date):
            return datetime(val.year, val.month, val.day).strftime("%Y-%m-%d")
    except Exception: pass
    s = str(val).strip()
    if not s or s in ("nan","None","NaT","NaN",""): return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if not (2020 <= y <= 2035): return None
        mo, d = (b, a) if a > 12 else (a, b)
        if 1 <= mo <= 12 and 1 <= d <= 31: return f"{y}-{mo:02d}-{d:02d}"
    try:
        dt = pd.to_datetime(s, dayfirst=False, errors="coerce")
        if pd.notna(dt) and 2020 <= dt.year <= 2035:
            return dt.strftime("%Y-%m-%d")
    except Exception: pass
    return None

def fmt_day(ds):
    d = datetime.strptime(ds, "%Y-%m-%d")
    return f"{d.strftime('%b')} {d.day}"

def fmt_date_range(s1, s2):
    try:
        a = datetime.strptime(s1, "%Y-%m-%d")
        b = datetime.strptime(s2, "%Y-%m-%d")
        al = f"{a.strftime('%b')} {a.day}"
        bl = str(b.day) if a.month == b.month else f"{b.strftime('%b')} {b.day}"
        return f"{al}–{bl}"
    except Exception:
        return f"{s1}–{s2}"

def dot_color(n, mn, ideal):
    """mn=5 (critical threshold), ideal=8 (daily goal).
    ≥ideal=green, (mn+1)–(ideal-1)=orange, ≤mn=red, 0=absent."""
    if not n:      return "#FEE2CC"   # absent
    if n >= ideal: return GRN          # ≥8 goal
    if n > mn:     return ORG          # 6-7 acceptable
    return RED                         # ≤5 critical

def badge_for(total, wmin, wideal):
    """wmin=25 (critical threshold), wideal=40 (weekly goal)."""
    if not total:       return "No data",    "#FEE2E2", RED
    if total >= wideal: return "Goal",       "#DCFCE7", GRN
    if total > wmin:    return "In Progress","#FEF9C3", "#CA8A04"
    return "Critical", "#FEE2E2", RED

# ─────────────────────────────────────────────────────────────────────────────
# XLSX PARSERS
# ─────────────────────────────────────────────────────────────────────────────
def _week_header(cell_val):
    """Return (week_num, date_range_str_or_None) from a cell like
    'Week 1' or 'Week 2\\nMay 11-15\\n'."""
    s = str(cell_val or "").strip()
    m = re.match(r"^Week\s+(\d+)(?:\s*[\r\n]+(.+?))?(?:\s*[\r\n].*)?$",
                 s, re.DOTALL)
    if m:
        return int(m.group(1)), (m.group(2).strip() if m.group(2) else None)
    return None, None

def _dates_from_range(dr_str, year):
    """Parse 'May 4-8' → ('2026-05-04','2026-05-08') or (None,None)."""
    if not dr_str: return None, None
    m = re.match(r"(\w+)\s+(\d+)-(\d+)", dr_str.strip())
    if m:
        mn = MONTH_MAP.get(m.group(1).lower())
        if mn:
            try:
                return (date(year, mn, int(m.group(2))).strftime("%Y-%m-%d"),
                        date(year, mn, int(m.group(3))).strftime("%Y-%m-%d"))
            except Exception: pass
    return None, None

def _calc_week_dates(year, month, week_num):
    """Calculate Mon–Fri of the Nth work week of a calendar month.
    Week 1 = the week whose Monday is the first Monday on or after the 1st.
    Verified: May 2026 W1=May 4-8, W2=May 11-15, W3=May 18-22, W4=May 25-29.
    """
    try:
        first       = date(year, month, 1)
        wd          = first.weekday()                  # 0=Mon … 6=Sun
        days_to_mon = (7 - wd) % 7                    # 0 when already Monday
        first_mon   = first + timedelta(days_to_mon)
        week_start  = first_mon + timedelta(weeks=(week_num - 1))
        week_end    = week_start + timedelta(4)        # Friday
        return week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d')
    except Exception:
        return None, None


def parse_summary_sheet(wb, sheet_name, month_key, year):
    """
    Parse one Summary-Month sheet.
    Layout confirmed from file:
      Col A  (idx 0): Week header ("Week N" or "Week N\\nMay 4-8")
      Col B  (idx 1): MCC region label / group name
      Col C  (idx 2): MCC quota
      Col E  (idx 4): MCC 'Generated' / 'Target Batch' label
      Col F  (idx 5): MCC generated count / target total
      Col H  (idx 7): CS region label / group name
      Col I  (idx 8): CS quota
      Col K  (idx 10): CS 'Generated' / 'Target Batch' label
      Col L  (idx 11): CS generated count / target total

    Returns list of week dicts:
      { week_num, month_key, label, date_range,
        start, end,
        mcc: { total, groups:[{label,countries,quota,generated}] },
        cs:  { total, groups:[...] } }
    """
    ws = wb[sheet_name]
    weeks = []
    cur = None
    mcc_on = mcc_done = False
    cs_on  = cs_done  = False

    for raw_row in ws.iter_rows(values_only=True):
        row = list(raw_row) + [None] * 15
        if not any(v is not None for v in row):
            continue

        # ── Week header (col A) ─────────────────────────────────────────────
        if row[0] is not None:
            wn, dr = _week_header(row[0])
            if wn is not None:
                if cur:
                    weeks.append(cur)
                s, e = _dates_from_range(dr, year)
                lbl = f"Week {wn}" + (f" · {dr.replace('-','–')}" if dr else "")
                cur = dict(
                    week_num=wn, month_key=month_key,
                    label=lbl, date_range=dr,
                    start=s, end=e,
                    mcc=dict(total=0, groups=[]),
                    cs =dict(total=0, groups=[]),
                )
                mcc_on = mcc_done = False
                cs_on  = cs_done  = False
                continue

        if cur is None:
            continue

        # ── "Meta del Batch:" starts data collection ─────────────────────────
        if str(row[1] or "").strip() == "Meta del Batch:":
            mcc_on = True
        if str(row[7] or "").strip() == "Meta del Batch:":
            cs_on = True

        # ── MCC side (cols 1-5) ──────────────────────────────────────────────
        if mcc_on and not mcc_done:
            lbl4 = str(row[4] or "").strip()
            if lbl4 == "Target Batch":
                try: cur["mcc"]["total"] = int(float(str(row[5])))
                except Exception: pass
                mcc_done = True
            elif lbl4 != "Remaining for Goal":
                name = str(row[1] or "").strip()
                if name and name != "Meta del Batch:" and row[2] is not None:
                    try:
                        quota = int(float(str(row[2])))
                        gen   = int(float(str(row[4]))) if row[4] is not None else 0
                        ctrs  = [c.strip() for c in re.split(r",\s*", name) if c.strip()]
                        cur["mcc"]["groups"].append(
                            dict(label=name, countries=ctrs, quota=quota, generated=gen))
                    except Exception:
                        pass

        # ── CS side (cols 7-11) ─────────────────────────────────────────────
        if cs_on and not cs_done:
            lbl10 = str(row[10] or "").strip()
            if lbl10 == "Target Batch":
                try: cur["cs"]["total"] = int(float(str(row[11])))
                except Exception: pass
                cs_done = True
            elif lbl10 != "Remaining for Goal":
                name = str(row[7] or "").strip()
                if name and name != "Meta del Batch:" and row[8] is not None:
                    try:
                        quota = int(float(str(row[8])))
                        gen   = int(float(str(row[10]))) if row[10] is not None else 0
                        ctrs  = [c.strip() for c in re.split(r",\s*", name) if c.strip()]
                        cur["cs"]["groups"].append(
                            dict(label=name, countries=ctrs, quota=quota, generated=gen))
                    except Exception:
                        pass

    if cur:
        weeks.append(cur)
    return weeks


def parse_data_sheet(wb, sheet_name, month_key):
    """
    Parse one data sheet (e.g. 'May').

    Column groups confirmed from file analysis:
      Header: [None, Date, CaseID, CaseName, Country, Investigator, QANotes, QA'd,
               None, Date, CaseID, ...] repeating every 8 cols offset by 9.
      Date positions: [1, 9, 17, 25, 33, 41, 49, 57]

    Groups alternate MCC / CS by index:
      group index 0 (dp=1)  → Week 1 MCC
      group index 1 (dp=9)  → Week 1 CS
      group index 2 (dp=17) → Week 2 MCC
      group index 3 (dp=25) → Week 2 CS  … etc.

    Week assignment is by column POSITION, not by date.
    Verified: group 0 = exactly 120 cases, matching Summary Week 1 MCC total.

    Returns list of case dicts:
      { date, case_id, country, investigator, week_num, region, month_key }
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = list(rows[0])
    date_positions = [i for i, v in enumerate(header)
                      if str(v or "").strip() == "Date"]

    records = []
    seen    = set()

    for g_idx, dp in enumerate(date_positions):
        week_num = (g_idx // 2) + 1
        region   = "MCC" if g_idx % 2 == 0 else "CS"

        for raw_row in rows[1:]:
            row = list(raw_row)
            if dp >= len(row):
                continue
            try:
                # Offsets: +0=Date, +1=CaseID, +2=CaseName, +3=Country,
                #          +4=Investigator, +5=QANotes, +6=QA'd
                # ALL of Date, CaseID, CaseName, Country, Investigator must be
                # present for a row to count as a generated case.
                date_val      = row[dp]   if dp   < len(row) else None
                case_id_raw   = row[dp+1] if dp+1 < len(row) else None
                case_name_raw = row[dp+2] if dp+2 < len(row) else None
                country_raw   = row[dp+3] if dp+3 < len(row) else None
                inv_raw       = row[dp+4] if dp+4 < len(row) else None
                qa_notes      = row[dp+5] if dp+5 < len(row) else None

                ds = parse_date_val(date_val)
                if not ds:
                    continue

                case_id = str(case_id_raw or "").strip().strip("\"'")
                if not case_id or case_id in ("nan","None",""):
                    continue

                # Case Name must be present (required field)
                case_name = str(case_name_raw or "").strip()
                if not case_name or case_name in ("nan","None",""):
                    continue

                country = norm_country(str(country_raw or ""))
                if not country:
                    continue

                investigator = str(inv_raw or "").strip()
                if not investigator or investigator in ("nan","None",""):
                    continue

                if is_disq(qa_notes):
                    continue

                key = f"{case_id}|{investigator}"
                if key in seen:
                    continue
                seen.add(key)

                records.append(dict(
                    date=ds, case_id=case_id,
                    country=country, investigator=investigator,
                    week_num=week_num, region=region, month_key=month_key,
                ))
            except Exception:
                continue

    return records


def load_xlsx(file_obj):
    """
    Load the uploaded workbook. Populates:
      st.session_state.all_weeks    – list of week dicts (all months, sorted)
      st.session_state.all_cases    – deduplicated case dicts
      st.session_state.current_week_idx / prev_week_idx
      st.session_state.rcfg         – updated from current week's Summary groups
    """
    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as e:
        st.error(f"Could not open file: {e}")
        return False

    year        = datetime.today().year
    all_weeks   = []
    all_cases   = []
    seen_keys   = set()

    for sname in wb.sheetnames:
        slo = sname.strip().lower()

        if slo.startswith("summary"):
            # e.g. "Summary-May" → month part = "may"
            month_part = slo.split("-", 1)[1].strip() if "-" in slo else ""
            month_num  = MONTH_MAP.get(month_part)
            if month_num is None:
                continue
            month_key = f"{year}-{month_num:02d}"
            try:
                weeks = parse_summary_sheet(wb, sname, month_key, year)
                all_weeks.extend(weeks)
            except Exception as ex:
                st.warning(f"⚠ Could not parse {sname}: {ex}")
        else:
            # Plain month sheet e.g. "May", "April"
            month_num = MONTH_MAP.get(slo)
            if month_num is None:
                continue
            month_key = f"{year}-{month_num:02d}"
            try:
                cases = parse_data_sheet(wb, sname, month_key)
                for c in cases:
                    key = f"{c['case_id']}|{c['investigator']}"
                    if key not in seen_keys:
                        seen_keys.add(key)
                        all_cases.append(c)
            except Exception as ex:
                st.warning(f"⚠ Could not parse {sname}: {ex}")

    wb.close()

    # Sort weeks chronologically
    all_weeks.sort(key=lambda w: (w["month_key"], w["week_num"]))

    # Fill missing date ranges using CALENDAR MATH (not case dates).
    # Months without explicit ranges in the Summary header (March, April) used
    # to get start/end from min/max of case dates, which can span months and
    # cause today's date to incorrectly match a past week.
    # _calc_week_dates() gives the exact Mon-Fri for week N of a month,
    # verified against all May Summary headers.
    for w in all_weeks:
        if not w["start"] or not w["end"]:
            y  = int(w["month_key"][:4])
            mo = int(w["month_key"][5:])
            s, e = _calc_week_dates(y, mo, w["week_num"])
            if s and e:
                w["start"] = s
                w["end"]   = e
                w["label"] = f"Week {w['week_num']} · {fmt_date_range(s, e)}"

    # Determine current week: find the week whose Mon-Fri span contains today.
    # Because all weeks now have calendar-correct start/end, each non-overlapping
    # Mon-Fri range will match at most one week.
    today_str = date.today().strftime("%Y-%m-%d")
    cur_idx   = None
    for i, w in enumerate(all_weeks):
        if w.get("start") and w.get("end"):
            if w["start"] <= today_str <= w["end"]:
                cur_idx = i
                break

    if cur_idx is None:
        # Today is not inside any week (e.g. file only covers past months).
        # Use the most recent past week.
        past = [i for i, w in enumerate(all_weeks)
                if w.get("end") and w["end"] < today_str]
        cur_idx = past[-1] if past else max(0, len(all_weeks) - 1)

    prev_idx = max(0, cur_idx - 1)

    # Update session state
    st.session_state.all_weeks        = all_weeks
    st.session_state.all_cases        = all_cases
    st.session_state.current_week_idx = cur_idx
    st.session_state.prev_week_idx    = prev_idx
    st.session_state.view             = "current"
    st.session_state.full_month_key   = all_weeks[cur_idx]["month_key"] if all_weeks else None

    # Rebuild rcfg from current week's Summary so sidebar shows live data
    if all_weeks and cur_idx < len(all_weeks):
        cw = all_weeks[cur_idx]
        new_rcfg = copy.deepcopy(DEFAULT_REGIONS)
        for rk, rkey in (("MCC","mcc"),("CS","cs")):
            if cw[rkey]["groups"]:
                new_rcfg[rk]["groups"] = [
                    dict(label=g["label"], countries=g["countries"], quota=g["quota"])
                    for g in cw[rkey]["groups"]
                ]
        st.session_state.rcfg = new_rcfg

    n = len(all_cases)
    nw = len(all_weeks)
    st.toast(f"✅ {n} cases loaded across {nw} weeks", icon="📊")
    return True

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  .stApp{{background-color:{BG}!important}}
  .main .block-container{{padding:1rem 2rem 2rem;max-width:1440px}}
  #MainMenu,footer,header{{visibility:hidden}}
  [data-testid="stVerticalBlockBorderWrapper"]{{border:1px solid {BORD}!important;border-radius:14px!important;background:{CARD}!important}}
  .sec-lbl{{font-size:10px;font-weight:700;color:{ORG};letter-spacing:.08em;text-transform:uppercase;margin-bottom:12px}}
  .pw{{height:7px;background:{ABSC};border-radius:4px;margin:4px 0}}
  .pf{{height:100%;border-radius:4px}}
  .hl{{display:flex;align-items:flex-start;gap:8px;margin-bottom:10px}}
  .hd{{width:7px;height:7px;border-radius:50%;margin-top:4px;flex-shrink:0;display:inline-block}}
  [data-testid="stButton"] button{{transition:background .15s,border-color .15s!important}}
  [data-testid="stButton"] button[kind="primary"]{{background:{ORG}!important;color:#fff!important;border:none!important}}
  [data-testid="stButton"] button[kind="primary"]:hover{{background:#EA6C0A!important}}
  [data-testid="stButton"] button[kind="secondary"]{{background:{CARD}!important;color:{TX}!important;border:1px solid {BORD}!important}}
  [data-testid="stButton"] button[kind="secondary"]:hover{{background:{OL}!important;border-color:{ORG}!important;color:{ORG}!important}}
  [data-testid="stButton"] button:disabled{{opacity:.35!important}}
  [data-testid="stFileUploader"]{{background:{CARD}!important;border:1px dashed {BORD}!important;border-radius:8px!important}}
  hr{{border-color:{BORD};margin:6px 0}}
  p,span,label,div{{color:{TX}}}
  [data-testid="stSelectbox"]>div>div{{background:{CARD}!important;border-color:{BORD}!important;color:{TX}!important}}
  [data-baseweb="select"]>div{{background:{CARD}!important;border-color:{BORD}!important}}
  [data-baseweb="select"] span,[data-baseweb="select"] div{{color:{TX}!important}}
  [data-baseweb="select"] svg{{fill:{TX2}!important}}
  [data-baseweb="popover"]{{background:{CARD}!important;border:1px solid {BORD}!important;border-radius:8px!important;box-shadow:none!important}}
  [data-baseweb="menu"],[data-baseweb="list"]{{background:{CARD}!important}}
  [role="option"]{{background:{CARD}!important;color:{TX}!important}}
  [role="option"]:hover,[role="option"][aria-selected="true"]{{background:{OL}!important;color:{ORG}!important}}
  [data-testid="stSidebar"]{{background:{CARD}!important;border-right:1px solid {BORD}}}
  [data-testid="stSidebar"] p,[data-testid="stSidebar"] span,[data-testid="stSidebar"] label{{color:{TX}!important}}
  [data-testid="stSidebar"] [data-baseweb="select"]>div{{background:{BG}!important;border-color:{BORD}!important}}
  [data-testid="stSidebar"] input{{background:{BG}!important;color:{TX}!important;border-color:{BORD}!important}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Country Configuration")
    st.caption("Changes apply instantly to all displayed data.")
    if st.button("↩ Reset to defaults", use_container_width=True):
        st.session_state.rcfg = copy.deepcopy(DEFAULT_REGIONS); st.rerun()
    st.markdown("---")
    assigned   = get_all_assigned()
    all_pool   = sorted(set(ALL_KNOWN_COUNTRIES) | set(assigned.keys()))
    unassigned = [c for c in all_pool if c not in assigned]
    for rk in ["MCC","CS"]:
        rc = st.session_state.rcfg[rk]
        st.markdown(f"### 🌎 {rc['name']}")
        st.caption(f"Total quota: **{total_quota(rk)}** cases")
        for gi, g in enumerate(list(rc["groups"])):
            with st.expander(f"📦 {g['label']} ({g['quota']})", expanded=False):
                nq = st.number_input("Quota", value=g["quota"], min_value=0, step=1,
                                     key=f"q_{rk}_{gi}")
                if nq != g["quota"]:
                    st.session_state.rcfg[rk]["groups"][gi]["quota"] = int(nq); st.rerun()
                nl = st.text_input("Name", value=g["label"], key=f"lbl_{rk}_{gi}")
                if nl != g["label"]:
                    st.session_state.rcfg[rk]["groups"][gi]["label"] = nl; st.rerun()
                st.markdown("**Countries:**")
                for ctr in list(g["countries"]):
                    ca, cb = st.columns([5,1]); ca.markdown(f"🌍 {ctr}")
                    if cb.button("✕", key=f"rm_{rk}_{gi}_{ctr}"):
                        st.session_state.rcfg[rk]["groups"][gi]["countries"].remove(ctr)
                        st.rerun()
                others  = [c for c,(r,_) in assigned.items() if r != rk]
                movable = sorted(set(unassigned + others))
                if movable:
                    pick = st.selectbox("Add / move", ["— select —"]+movable, key=f"add_{rk}_{gi}")
                    if pick and pick != "— select —":
                        for rk2 in st.session_state.rcfg:
                            for g2 in st.session_state.rcfg[rk2]["groups"]:
                                if pick in g2["countries"]: g2["countries"].remove(pick)
                        st.session_state.rcfg[rk]["groups"][gi]["countries"].append(pick)
                        st.rerun()
                cust = st.text_input("Add unlisted", placeholder="Country name",
                                     key=f"cust_{rk}_{gi}")
                if cust:
                    cn = cust.strip().title()
                    if cn and cn not in g["countries"]:
                        for rk2 in st.session_state.rcfg:
                            for g2 in st.session_state.rcfg[rk2]["groups"]:
                                if cn in g2["countries"]: g2["countries"].remove(cn)
                        st.session_state.rcfg[rk]["groups"][gi]["countries"].append(cn)
                        st.rerun()
                if not g["countries"]:
                    if st.button("🗑 Delete group", key=f"del_{rk}_{gi}"):
                        st.session_state.rcfg[rk]["groups"].pop(gi); st.rerun()
        with st.expander("➕ New group", expanded=False):
            nn = st.text_input("Name", placeholder="e.g. Venezuela", key=f"ng_{rk}")
            nq = st.number_input("Quota", value=5, min_value=0, key=f"ngq_{rk}")
            if st.button(f"Add to {rk}", key=f"ngb_{rk}") and nn:
                st.session_state.rcfg[rk]["groups"].append(
                    {"label":nn,"countries":[],"quota":int(nq)}); st.rerun()
        st.markdown("---")
    if unassigned:
        st.markdown("### ⚠️ Unassigned Countries")
        for c in unassigned: st.markdown(f"- {c}")

# ─────────────────────────────────────────────────────────────────────────────
# TOP BAR
# ─────────────────────────────────────────────────────────────────────────────
cfg   = st.session_state.rcfg[st.session_state.tab]
pills = region_pills(st.session_state.tab)
cl, cm, cs_, cr, ct = st.columns([1.4, 1, 1, 4.5, 0.5])

with cl:
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:8px;padding:5px 0">
      <div style="width:30px;height:30px;background:{ORG};border-radius:7px;display:flex;
                  align-items:center;justify-content:center;color:white;font-weight:800;font-size:14px">R</div>
      <div>
        <div style="font-weight:700;font-size:13px;color:{TX}">ruvixx</div>
        <div style="font-size:8px;color:{TX2};letter-spacing:.06em;text-transform:uppercase">Case Investigation</div>
      </div>
    </div>""", unsafe_allow_html=True)

with cm:
    if st.button("México CC", key="btn_mcc",
                 type="primary" if st.session_state.tab == "MCC" else "secondary",
                 use_container_width=True):
        st.session_state.tab = "MCC"; st.rerun()

with cs_:
    if st.button("Cono Sur", key="btn_cs",
                 type="primary" if st.session_state.tab == "CS" else "secondary",
                 use_container_width=True):
        st.session_state.tab = "CS"; st.rerun()

with cr:
    ph = "".join(
        f'<span style="font-size:9px;font-weight:700;background:{ORG};color:white;'
        f'border-radius:3px;padding:1px 4px;margin:0 1px">{p}</span>' for p in pills)
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:6px;background:{OL};border:1px solid {OB};
                border-radius:8px;padding:6px 10px;font-size:12px;color:#92400E;
                font-weight:500;flex-wrap:wrap">
      <span style="width:7px;height:7px;border-radius:50%;background:{ORG};
                   display:inline-block;flex-shrink:0"></span>
      {cfg["name"]} · {cfg["contact"]} &nbsp;{ph}
    </div>""", unsafe_allow_html=True)

with ct:
    if st.button("🌙" if not dark else "☀️", key="theme_btn", use_container_width=True):
        st.session_state.dark = not dark; st.rerun()

st.markdown(f'<hr style="border-color:{BORD}">', unsafe_allow_html=True)

# ── Data source status strip ─────────────────────────────────────────────────
_xlsx_ok  = st.session_state.xlsx_fetch_ok
_batch_ok = st.session_state.batch_fetch_ok
_xlsx_ts  = st.session_state.xlsx_fetch_ts or "—"
_xlsx_err = st.session_state.xlsx_fetch_err

_xlsx_dot  = f'<span style="color:{GRN}">●</span>' if _xlsx_ok  else f'<span style="color:{RED}">●</span>'
_batch_dot = f'<span style="color:{GRN}">●</span>' if _batch_ok else f'<span style="color:{TX2}">●</span>'
_xlsx_lbl  = (f'<b>LATAM xlsx</b> · fetched {_xlsx_ts}'
              if _xlsx_ok
              else f'<b>LATAM xlsx</b> · <span style="color:{RED}">not loaded</span>'
                   + (f' — {_xlsx_err}' if _xlsx_err else ''))
_batch_lbl = ('<b>Batch history</b> · loaded' if _batch_ok
              else f'<b>Batch history</b> · <span style="color:{TX2}">unavailable</span>')

st.markdown(
    f'<div style="display:flex;align-items:center;gap:20px;padding:5px 2px 8px;'
    f'font-size:11px;color:{TX2}">'
    f'<span>{_xlsx_dot} {_xlsx_lbl}</span>'
    f'<span style="color:{BORD}">|</span>'
    f'<span>{_batch_dot} {_batch_lbl}</span>'
    f'</div>',
    unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# AUTO-FETCH ON STARTUP — both xlsx and batch history
# ─────────────────────────────────────────────────────────────────────────────
if not st.session_state.xlsx_fetch_ok:
    with st.spinner("Loading data from GitHub…"):
        fetch_xlsx_from_github(force=False)
        fetch_batch_history()

# ─────────────────────────────────────────────────────────────────────────────
# CONTROLS  —  REFRESH  +  VIEW SELECTOR
# ─────────────────────────────────────────────────────────────────────────────
all_weeks = st.session_state.all_weeks
cur_idx   = st.session_state.current_week_idx
prev_idx  = st.session_state.prev_week_idx
view      = st.session_state.view
has_weeks = bool(all_weeks)
has_prev  = has_weeks and (prev_idx != cur_idx)

c_ref, c_week, c_fm, c_msel, c_st = st.columns([1.4, 2.3, 1.3, 1.9, 2.9])

with c_ref:
    if st.button("🔄 Refresh", key="btn_refresh",
                 type="secondary", use_container_width=True,
                 help="Re-fetch xlsx and batch history from GitHub"):
        with st.spinner("Refreshing…"):
            ok_x = fetch_xlsx_from_github(force=True)
            ok_b = fetch_batch_history()
        if ok_x:
            st.toast("✅ Data refreshed from GitHub", icon="📊")
            st.rerun()
        else:
            st.error(st.session_state.xlsx_fetch_err or "Refresh failed.")

# ── Week selector dropdown (Current / Previous) ──────────────────────────────
# IMPORTANT: do NOT use key= on this selectbox. With a key, Streamlit caches the
# widget's own value and ignores index= on reruns — causing stale view state.
# Without a key, index= fully controls the displayed and returned value each run.
with c_week:
    if has_weeks:
        def _wk_label(idx, prefix):
            w = all_weeks[idx]
            if w.get("start") and w.get("end"):
                return f"{prefix} {fmt_date_range(w['start'], w['end'])}"
            return f"{prefix} Week {w['week_num']}"

        week_options      = []
        week_option_views = []
        week_options.append(_wk_label(cur_idx, "★ Current:"))
        week_option_views.append("current")
        if has_prev:
            week_options.append(_wk_label(prev_idx, "← Previous:"))
            week_option_views.append("prev")

        # index= drives the DISPLAYED selection on every run — no key caching
        _cur_view = view if view in week_option_views else "current"
        wk_sel_idx = week_option_views.index(_cur_view)

        wk_choice = st.selectbox(
            "Week", week_options, index=wk_sel_idx,
            label_visibility="collapsed",
            disabled=(view == "full_month"),
        )
        # Detect user change and update session state immediately
        if view != "full_month":
            chosen_view = week_option_views[week_options.index(wk_choice)]
            if chosen_view != st.session_state.view:
                st.session_state.view = chosen_view
                st.rerun()
    else:
        st.markdown(
            f'<div style="font-size:11px;color:{TX2};padding-top:8px">'
            f'Waiting for data…</div>', unsafe_allow_html=True)

# ── Full Month toggle ─────────────────────────────────────────────────────────
with c_fm:
    if st.button(
        "📅 Full Month",
        key="btn_month",
        type="primary" if view == "full_month" else "secondary",
        use_container_width=True,
        disabled=not has_weeks,
    ):
        if view == "full_month":
            # Toggle off → back to current week
            st.session_state.view = "current"
        else:
            st.session_state.view = "full_month"
            # Default full_month_key to current month
            if not st.session_state.full_month_key and has_weeks:
                st.session_state.full_month_key = all_weeks[cur_idx]["month_key"]
        st.rerun()

# ── Month selector (visible only in full_month mode) ─────────────────────────
with c_msel:
    if view == "full_month" and has_weeks:
        # Build available months: current + previous (based on today)
        today = date.today()
        cur_mk  = f"{today.year}-{today.month:02d}"
        # Previous calendar month
        prev_mo_date = date(today.year, today.month, 1) - timedelta(1)
        prev_mk = f"{prev_mo_date.year}-{prev_mo_date.month:02d}"

        # Only offer months that actually have data in the workbook
        avail_month_keys = sorted(set(w["month_key"] for w in all_weeks))
        month_opts = [mk for mk in [cur_mk, prev_mk] if mk in avail_month_keys]
        if not month_opts:
            month_opts = avail_month_keys[-2:] if len(avail_month_keys) >= 2 else avail_month_keys

        month_labels = {mk: datetime.strptime(mk, "%Y-%m").strftime("%B %Y")
                        for mk in month_opts}

        # Ensure full_month_key is valid
        if st.session_state.full_month_key not in month_opts:
            st.session_state.full_month_key = month_opts[-1]

        mo_default = month_opts.index(st.session_state.full_month_key)
        mo_choice  = st.selectbox(
            "Month", options=month_opts,
            format_func=lambda mk: month_labels[mk],
            index=mo_default,
            label_visibility="collapsed",
            key="month_dropdown",
        )
        if mo_choice != st.session_state.full_month_key:
            st.session_state.full_month_key = mo_choice; st.rerun()
    else:
        st.markdown("", unsafe_allow_html=True)  # empty placeholder

with c_st:
    if not has_weeks:
        _err = st.session_state.xlsx_fetch_err
        if _err:
            st.markdown(
                f'<div style="font-size:11px;color:{RED};padding-top:4px;line-height:1.6">'
                f'⚠️ {_err}</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div style="font-size:11px;color:{TX2};padding-top:8px">'
                f'⏳ Loading…</div>', unsafe_allow_html=True)
    else:
        df_all = pd.DataFrame(st.session_state.all_cases)
        nm = len(df_all[df_all["region"]=="MCC"]) if not df_all.empty else 0
        nc = len(df_all[df_all["region"]=="CS"])  if not df_all.empty else 0
        st.markdown(
            f'<div style="font-size:11px;color:{TX2};padding-top:2px;line-height:1.7">'
            f'{len(st.session_state.all_cases)} cases · {len(all_weeks)} weeks<br>'
            f'MCC <b style="color:{ORG}">{nm}</b> &nbsp;·&nbsp; '
            f'CS <b style="color:{ORG}">{nc}</b>'
            f'</div>', unsafe_allow_html=True)

# ── Empty state — GitHub fetch failed, offer manual upload as fallback ────────
if not has_weeks:
    _err       = st.session_state.xlsx_fetch_err or ""
    _file_missing = "not found" in _err.lower() or "404" in _err
    _no_token     = "GITHUB_TOKEN" in _err

    st.markdown(f'<hr style="border-color:{BORD};margin:0 0 24px">', unsafe_allow_html=True)

    col_info, col_upload = st.columns([1.2, 1])

    with col_info:
        st.markdown(f"""
        <div style="padding:16px 0">
          <div style="font-size:18px;font-weight:700;color:{TX};margin-bottom:8px">
            📊 Data not loaded yet</div>
          <div style="font-size:13px;color:{TX2};line-height:1.8;margin-bottom:12px">
            The dashboard auto-fetches the xlsx from GitHub.<br>
            If the file is not there yet, upload it once below to bootstrap.<br>
            After the first upload the Apps Script will keep it in sync.
          </div>
        """, unsafe_allow_html=True)

        if _err:
            _err_color = RED
            st.markdown(
                f'<div style="font-size:12px;color:{_err_color};background:#FEE2E2;'
                f'border-radius:8px;padding:10px 14px;margin-bottom:10px">'
                f'⚠️ <b>GitHub fetch error:</b><br>{_err}</div>',
                unsafe_allow_html=True)

        if _file_missing:
            st.markdown(f"""
            <div style="font-size:12px;color:{TX2};background:{OL};border:1px solid {OB};
                        border-radius:8px;padding:10px 14px;line-height:1.7">
              <b>To push the file from Apps Script (one-time):</b><br>
              1. Open the Google Sheet → Extensions → Apps Script<br>
              2. Select <code>pushXlsxToGitHub</code> from the function dropdown<br>
              3. Click <b>Run</b> — it will create the file in the repo<br>
              4. Come back and click 🔄 Refresh above
            </div>""", unsafe_allow_html=True)
        elif _no_token:
            st.markdown(f"""
            <div style="font-size:12px;color:{TX2};background:{OL};border:1px solid {OB};
                        border-radius:8px;padding:10px 14px;line-height:1.7">
              <b>To set the token in Streamlit secrets:</b><br>
              App settings → Secrets → add:<br>
              <code>GITHUB_TOKEN = "ghp_..."</code>
            </div>""", unsafe_allow_html=True)

        st.markdown("</div>", unsafe_allow_html=True)

    with col_upload:
        st.markdown(
            f'<div style="font-size:12px;font-weight:700;color:{TX};margin-bottom:6px">'
            f'⬆️ Manual upload (bootstrap)</div>',
            unsafe_allow_html=True)
        st.caption("Upload the xlsx once. After that, GitHub auto-sync takes over.")
        _fallback = st.file_uploader(
            "Upload LATAM xlsx", type=["xlsx"],
            label_visibility="collapsed",
            key="xlsx_fallback_upload",
            help="Bootstrap: upload the xlsx manually while GitHub sync is not yet set up",
        )
        if _fallback is not None:
            with st.spinner("Parsing file…"):
                _ok = load_xlsx(_fallback)
            if _ok:
                st.session_state.xlsx_fetch_ok  = True
                st.session_state.xlsx_fetch_ts  = datetime.now().strftime("%b %d, %H:%M")
                st.session_state.xlsx_fetch_err = None
                st.toast("✅ File loaded — dashboard ready", icon="📊")
                st.rerun()
            else:
                st.error("Could not parse the xlsx. Check the file and try again.")

    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# COMPUTE
# ─────────────────────────────────────────────────────────────────────────────
tab           = st.session_state.tab
cfg           = st.session_state.rcfg[tab]
view          = st.session_state.view
view_is_month = (view == "full_month")

# ── Determine selected week and month context ─────────────────────────────────
# Week views: anchor on the selected week's calendar period
# Full month: use the user-chosen month key
sel_idx  = prev_idx if view == "prev" else cur_idx
sel_week = all_weeks[sel_idx]

if view_is_month:
    cur_month_key = st.session_state.full_month_key or all_weeks[cur_idx]["month_key"]
else:
    cur_month_key = sel_week["month_key"]

month_weeks = [w for w in all_weeks if w["month_key"] == cur_month_key]

# ── Calendar boundaries ───────────────────────────────────────────────────────
# These are always derived from the calendar, never from the xlsx Summary header.
# Summary headers may contain batch date ranges (e.g. "Apr 29-May 15") that
# span multiple calendar weeks and would produce wrong filter windows.
import calendar as _cal
_y, _m = int(cur_month_key[:4]), int(cur_month_key[5:])
month_start_str = f"{_y}-{_m:02d}-01"
month_end_str   = f"{_y}-{_m:02d}-{_cal.monthrange(_y, _m)[1]:02d}"

# Week boundary: ALWAYS computed from calendar math for the exact Mon–Fri 5-day
# window. sel_week["week_num"] and sel_week["month_key"] are the only trusted
# fields; sel_week["start"]/["end"] may reflect batch periods from the Summary
# header which can span multiple weeks and must NOT be used here.
_wk_year  = int(sel_week["month_key"][:4])
_wk_month = int(sel_week["month_key"][5:])
week_start_str, week_end_str = _calc_week_dates(_wk_year, _wk_month, sel_week["week_num"])
# Safety fallback (should never occur given _calc_week_dates is always valid)
if not week_start_str or not week_end_str:
    week_start_str = sel_week.get("start") or ""
    week_end_str   = sel_week.get("end")   or ""

# ── Build region-filtered base pool ──────────────────────────────────────────
cases_df = pd.DataFrame(st.session_state.all_cases) if st.session_state.all_cases else EMPTY_DF
has_data = not cases_df.empty

# STEP 1: column-position region tag (set during xlsx parsing by column group)
r_data = cases_df[cases_df["region"] == tab].copy() if has_data else pd.DataFrame()

# STEP 2: country-membership validation.
# Use DEFAULT_REGIONS (not rcfg) as the authoritative country list.
# rcfg is dynamically updated from the current week's Summary groups and may
# drop countries when the Summary changes week-to-week, causing valid cases from
# other weeks to be silently excluded. DEFAULT_REGIONS always holds the full
# canonical list for each region.
_default_countries = set(
    c for g in DEFAULT_REGIONS[tab]["groups"] for c in g["countries"]
)
# Also include any countries present in all Summary groups across all weeks
# (e.g. countries added by a newer batch that aren't in DEFAULT_REGIONS yet)
_summary_countries = set()
for _w in all_weeks:
    _rk = "mcc" if tab == "MCC" else "cs"
    for _g in _w.get(_rk, {}).get("groups", []):
        for _c in _g.get("countries", []):
            _summary_countries.add(_c)

region_countries = _default_countries | _summary_countries

if not r_data.empty:
    r_data = r_data[r_data["country"].isin(region_countries)]

# ── w_data: STRICT calendar-date filter ──────────────────────────────────────
# RULE: the view label (e.g. "May 4–8") defines the EXACT date window shown.
# Cases worked before/after that window do NOT belong to this view, even if
# they are in the same column group in the xlsx.
if view_is_month:
    # Full month: all cases with dates inside the calendar month (May 1–31)
    w_data = r_data[
        (r_data["date"] >= month_start_str) &
        (r_data["date"] <= month_end_str)
    ] if not r_data.empty else pd.DataFrame()
else:
    # Week view: ONLY cases dated within the calendar week (Mon–Fri)
    w_data = r_data[
        (r_data["date"] >= week_start_str) &
        (r_data["date"] <= week_end_str)
    ] if not r_data.empty and week_start_str and week_end_str else pd.DataFrame()

# m_data: same month boundary, used for "month total" shown on week-view cards
m_data = r_data[
    (r_data["date"] >= month_start_str) &
    (r_data["date"] <= month_end_str)
] if not r_data.empty else pd.DataFrame()

# ── Summary quota (target only — not used for generated count) ───────────────
rkey = "mcc" if tab == "MCC" else "cs"

if view_is_month:
    tq = sum(w[rkey]["total"] for w in month_weeks)
    grp_map = {}
    for w in month_weeks:
        for g in w[rkey]["groups"]:
            lbl = g["label"]
            if lbl not in grp_map:
                grp_map[lbl] = dict(label=lbl, countries=g["countries"], quota=0)
            grp_map[lbl]["quota"] += g["quota"]
    summary_groups = list(grp_map.values())
    view_label = datetime.strptime(cur_month_key, "%Y-%m").strftime("%B %Y")
else:
    tq = sel_week[rkey]["total"]
    summary_groups = [
        dict(label=g["label"], countries=g["countries"], quota=g["quota"])
        for g in sel_week[rkey]["groups"]
    ]
    view_label = sel_week["label"]

# ── Batch-history exclusion ───────────────────────────────────────────────────
# delivered_ids: set of case IDs already shipped in a previous batch for this region.
# new_w_data: valid cases in the current view that are NOT yet delivered.
# These drive the batch progress metrics (donut, group breakdown, highlights).
# Investigator cards use w_data (all valid cases — their full productivity).
delivered_ids   = st.session_state.delivered_ids.get(tab, set())
has_batch_data  = st.session_state.batch_fetch_ok

new_w_data = (w_data[~w_data["case_id"].isin(delivered_ids)]
              if not w_data.empty and delivered_ids
              else w_data.copy())

# ── Batch progress totals (data-sheet driven, batch-filtered) ─────────────────
# total_new   = new cases not in any previous batch  → current batch progress
# total_deliv = cases from the current view already in a previous batch
# total_all   = all valid cases this period (investigator productivity)
total_new   = len(new_w_data) if not new_w_data.empty else 0
total_deliv = (len(w_data) - total_new) if not w_data.empty else 0
total_all   = len(w_data) if not w_data.empty else 0

# For the headline donut: progress = new cases / target quota
total = total_new
gap   = max(0, tq - total)
pct   = round(total / tq * 100) if tq else 0

# ── Group breakdown: new cases per country group ──────────────────────────────
groups = []
for g in summary_groups:
    g_countries = set(g["countries"])
    if not new_w_data.empty:
        g_new = len(new_w_data[new_w_data["country"].isin(g_countries)])
    else:
        g_new = 0
    groups.append(dict(
        label=g["label"], quota=g["quota"], eff_quota=g["quota"], done=g_new,
    ))

# ── Investigator stats from data sheet ───────────────────────────────────────
# For full month: compute per-week breakdown for each investigator.
# For week view: compute per-day breakdown within the calendar week.

# Build a mapping of week_num → (start, end) for the current month
_wk_ranges = {
    w["week_num"]: (w["start"], w["end"])
    for w in month_weeks
    if w.get("start") and w.get("end")
}

invs = []
if not w_data.empty:
    for inv_name, grp in sorted(w_data.groupby("investigator"),
                                key=lambda x: -len(x[1])):
        month_total = (len(m_data[m_data["investigator"] == inv_name])
                       if not m_data.empty else 0)
        by_day = grp.groupby("date").size().to_dict()

        # Weekly breakdown for full-month view:
        # For each week of the month, count dates within that week's calendar range
        by_week = {}
        for wn, (ws, we) in sorted(_wk_ranges.items()):
            wk_count = len(grp[(grp["date"] >= ws) & (grp["date"] <= we)])
            by_week[wn] = wk_count

        invs.append(dict(
            name=inv_name, total=len(grp), month_total=month_total,
            by_day=by_day, by_week=by_week,
            support=(inv_name in cfg.get("support", [])),
        ))

by_country = (w_data.groupby("country").size()
              .sort_values(ascending=False).to_dict()
              if not w_data.empty else {})

by_inv_stat = [dict(name=i["name"], total=i["total"],
                    pct=round(i["total"]/len(w_data)*100) if not w_data.empty else 0,
                    support=i["support"])
               for i in invs]

# ── w_days / w_weeks: display grid for production chart and daily bars ────────
_DAY = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

def _build_w_days(start_str, end_str):
    """Mon–Fri days between start and end (inclusive), with per-day case counts."""
    days = []
    if not start_str or not end_str:
        return days
    cur = datetime.strptime(start_str, "%Y-%m-%d")
    end = datetime.strptime(end_str,   "%Y-%m-%d")
    while cur <= end:
        if cur.weekday() < 5:
            ds = cur.strftime("%Y-%m-%d")
            n  = len(w_data[w_data["date"] == ds]) if not w_data.empty else 0
            days.append(dict(ds=ds, day=_DAY[cur.weekday()],
                             label=fmt_day(ds), total=n))
        cur += timedelta(1)
    return days

if view_is_month:
    # Full month: build w_days covering Mon–Fri of all weeks in the month
    starts = [w["start"] for w in month_weeks if w.get("start")]
    ends   = [w["end"]   for w in month_weeks if w.get("end")]
    w_days = _build_w_days(min(starts), max(ends)) if starts and ends else []

    # Also build weekly aggregates for the production chart and investigator bars
    w_weeks = []
    for w in sorted(month_weeks, key=lambda x: x["week_num"]):
        if not w.get("start") or not w.get("end"):
            continue
        wk_total = len(w_data[
            (w_data["date"] >= w["start"]) & (w_data["date"] <= w["end"])
        ]) if not w_data.empty else 0
        w_weeks.append(dict(
            week_num=w["week_num"],
            label=f"W{w['week_num']}",
            detail=fmt_date_range(w["start"], w["end"]),
            total=wk_total,
        ))
else:
    # Week view: exactly Mon–Fri of the selected calendar week (always 5 days)
    w_days  = _build_w_days(week_start_str, week_end_str)
    w_weeks = []  # not used in week view

# ─────────────────────────────────────────────────────────────────────────────
# METRIC ROW
# ─────────────────────────────────────────────────────────────────────────────
period_lbl = "Monthly quota" if view_is_month else "Weekly quota"

# Build delivered-batch indicator (shown only when history is loaded)
_deliv_tag = ""
if has_batch_data and total_deliv > 0:
    _deliv_tag = (f'<div style="font-size:9px;color:{TX2};margin-top:3px">'
                  f'<span style="color:{GRN}">+{total_new} new</span>'
                  f' · <span style="color:{TX2}">{total_deliv} prev batch</span></div>')
elif has_batch_data:
    _deliv_tag = (f'<div style="font-size:9px;color:{GRN};margin-top:3px">'
                  f'All new · no overlap</div>')

st.markdown(f"""
<div style="display:flex;justify-content:flex-end;align-items:center;
            gap:28px;padding:8px 0 12px">
  <div style="text-align:center">
    <div style="font-size:22px;font-weight:800;color:{ORG};line-height:1">{total_all}</div>
    <div style="font-size:9px;color:{TX2};text-transform:uppercase;letter-spacing:.06em">Cases Generated</div>
    {_deliv_tag}
  </div>
  <div style="text-align:center">
    <div style="font-size:22px;font-weight:800;color:{ORG};line-height:1">{gap}</div>
    <div style="font-size:9px;color:{TX2};text-transform:uppercase;letter-spacing:.06em">Batch Gap</div>
  </div>
  <div style="text-align:center">
    <div style="font-size:22px;font-weight:800;color:{ORG};line-height:1">{pct}%</div>
    <div style="font-size:9px;color:{TX2};text-transform:uppercase;letter-spacing:.06em">Batch Progress</div>
  </div>
  <div style="text-align:right">
    <div style="font-size:12px;font-weight:700;color:{TX}">Trimble LATAM</div>
    <div style="font-size:10px;color:{TX2}">{view_label}</div>
  </div>
</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# BATCH HISTORY PANEL
# Shows previously delivered batches for this region so investigators can see
# what has already been shipped and what counts as new progress.
# ─────────────────────────────────────────────────────────────────────────────
_batches, _total_delivered = batch_history_summary(tab)
if _batches:
    with st.expander(
        f"📦 Batch History — {cfg['name']} · "
        f"{len(_batches)} batch{'es' if len(_batches)!=1 else ''} · "
        f"{_total_delivered} total delivered",
        expanded=False,
    ):
        _cols = st.columns(min(4, len(_batches)))
        for _bi, _b in enumerate(_batches[:8]):
            with _cols[_bi % len(_cols)]:
                _bn   = _b.get("batch_number", "?")
                _dd   = _b.get("delivery_date", "—")
                _tc   = _b.get("total_cases", 0)
                _prof = _b.get("profile", "")
                st.markdown(f"""
                <div style="background:{CARD};border:1px solid {BORD};border-radius:10px;
                            padding:10px 12px;margin-bottom:6px">
                  <div style="font-size:11px;font-weight:700;color:{ORG}">Batch #{_bn}</div>
                  <div style="font-size:10px;color:{TX2};margin-top:2px">📅 {_dd}</div>
                  <div style="font-size:13px;font-weight:700;color:{TX};margin-top:4px">
                    {_tc} cases</div>
                  <div style="font-size:9px;color:{TX2};margin-top:2px;
                              white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
                    {_prof}</div>
                </div>""", unsafe_allow_html=True)
        if not has_batch_data:
            st.caption("⚠️ Batch history not loaded — add GITHUB_TOKEN to Streamlit secrets.")
elif not has_batch_data:
    st.markdown(
        f'<div style="font-size:11px;color:{TX2};padding:4px 0 8px">'
        f'ℹ️ Batch history unavailable — GITHUB_TOKEN not set. '
        f'New case counts reflect all generated cases.</div>',
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# MAIN ROW
# ─────────────────────────────────────────────────────────────────────────────
mc1, mc2, mc3 = st.columns([1.5, 3, 2])

with mc1:
    with st.container(border=True):
        fig_g = go.Figure(go.Pie(
            values=[max(total, 0.0001), max(gap, 0.0001)],
            hole=0.72, sort=False, textinfo="none", hoverinfo="none",
            marker_colors=[ORG, ABSC], showlegend=False,
        ))
        # total here = total_new (batch-filtered); total_all = all generated
        _inner_top   = f"<b>{total}</b>" if not has_batch_data else f"<b>{total}</b>"
        _inner_label = "/ " + str(tq) + " quota"
        for txt, yp, sz, col in [
            (_inner_top,    0.57, 26, TX),
            (_inner_label,  0.44, 10, TX2),
            (f"<b>{pct}%</b>", 0.30, 14, ORG),
            (period_lbl.upper(), 0.16, 8, TX2),
        ]:
            fig_g.add_annotation(text=txt, x=0.5, y=yp, showarrow=False,
                                  font=dict(size=sz, color=col))
        fig_g.update_layout(margin=dict(t=5,b=5,l=5,r=5), height=200,
                             paper_bgcolor="rgba(0,0,0,0)",
                             plot_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig_g, use_container_width=True,
                        config={"displayModeBar": False})
        q_lbl = "Monthly quota" if view_is_month else "Weekly quota"
        _dn = f" · {total_deliv} prev batch" if has_batch_data and total_deliv > 0 else ""
        st.markdown(
            f'<p style="text-align:center;font-size:11px;font-weight:700;'
            f'color:{ORG};margin-top:-20px">{gap} remaining · {total} new{_dn}</p>',
            unsafe_allow_html=True)

with mc2:
    with st.container(border=True):
        st.markdown('<div class="sec-lbl">Batch Quota · Group Breakdown</div>',
                    unsafe_allow_html=True)
        for g in groups:
            dq   = g["eff_quota"]
            left = max(0, dq - g["done"])
            bp   = min(100, g["done"] / dq * 100) if dq else 0
            bc   = GRN if left == 0 else (ORG if g["done"]/max(dq,1) >= 0.6 else RED)
            lbl  = "✓ done" if left == 0 else f"{left} left"
            st.markdown(f"""
            <div style="margin-bottom:9px">
              <div style="display:flex;justify-content:space-between;
                          font-size:12px;margin-bottom:3px">
                <span style="color:{TX}">{g['label']}</span>
                <span style="color:{TX2}">{g['done']}/{dq}
                  <span style="font-weight:700;color:{bc}">{lbl}</span></span>
              </div>
              <div class="pw"><div class="pf" style="width:{bp:.1f}%;background:{bc}">
              </div></div>
            </div>""", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="margin-top:10px;padding-top:8px;border-top:1px solid {BORD};
                    display:flex;justify-content:space-between;
                    font-size:11px;color:{TX2}">
          <span>Target Batch <b style="color:{TX}">{tq}</b></span>
          <span>Remaining <b style="color:{ORG}">{gap}</b></span>
        </div>""", unsafe_allow_html=True)

with mc3:
    with st.container(border=True):
        st.markdown('<div class="sec-lbl">⚡ Key Highlights</div>',
                    unsafe_allow_html=True)
        hl = [{"c": ORG, "t": "Current batch progress",
               "s": f"{total}/{tq} new cases — {gap} remaining"}]
        if has_batch_data and total_deliv > 0:
            hl.insert(0, {"c": TX2, "t": "Previously delivered",
                          "s": f"{total_deliv} cases already in past batches"})
        for g in groups:
            left = max(0, g["eff_quota"] - g["done"])
            hc   = GRN if left==0 else (ORG if g["done"]/max(g["eff_quota"],1) >= 0.6 else RED)
            hl.append({"c": hc, "t": g["label"],
                       "s": f"{g['done']}/{g['eff_quota']} — "
                            f"{'All complete ✓' if left==0 else f'{left} left'}"})
        if invs:
            tp = round(invs[0]["total"]/len(w_data)*100) if not w_data.empty else 0
            hl.append({"c": ORG, "t": "Top investigator",
                       "s": f"{invs[0]['name']} · {invs[0]['total']} ({tp}%)"})
        for h in hl[:9]:
            st.markdown(f"""
            <div class="hl">
              <span class="hd" style="background:{h['c']}"></span>
              <div>
                <div style="font-size:11px;color:{TX2}">{h['t']}</div>
                <div style="font-size:11px;font-weight:700;color:{h['c']}">{h['s']}</div>
              </div>
            </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# INVESTIGATOR CARDS
# ─────────────────────────────────────────────────────────────────────────────
if view_is_month:
    card_period = datetime.strptime(cur_month_key, "%Y-%m").strftime("%B %Y")
elif week_start_str and week_end_str:
    card_period = f"{sel_week['label'].split('·')[0].strip()} · {fmt_date_range(week_start_str, week_end_str)}"
else:
    card_period = sel_week["label"]
st.markdown(f"""
<div style="font-size:10px;font-weight:700;color:{TX2};letter-spacing:.07em;
            text-transform:uppercase;margin-bottom:8px">
  👤 Investigator Performance
  &nbsp;·&nbsp; Daily: ≥{cfg['daily_ideal']} goal
  <span style="color:{ORG}">· 6–{cfg['daily_ideal']-1} acceptable</span>
  <span style="color:{RED}">· ≤{cfg['daily_min']} critical</span>
  &nbsp;·&nbsp; Weekly: ≥{cfg['weekly_ideal']} goal
  <span style="color:{RED}">· ≤{cfg['weekly_min']} critical</span>
  &nbsp;·&nbsp; {card_period}
  <span style="font-size:10px;font-weight:400;background:{OL};padding:2px 8px;
               border-radius:20px;border:1px dashed {OB};margin-left:8px">
    click card to expand ↓</span>
</div>""", unsafe_allow_html=True)

n_weeks = len(month_weeks) if view_is_month else 1
w_min   = cfg["weekly_min"]   * n_weeks
w_ideal = cfg["weekly_ideal"] * n_weeks
month_str = datetime.strptime(cur_month_key, "%Y-%m").strftime("%B %Y")


def make_card(inv):
    bl, bb, bc = (("Support","#F3F4F6","#6B7280") if inv["support"]
                  else badge_for(inv["total"], w_min, w_ideal))
    wk_pct = min(100, inv["total"] / w_ideal * 100) if w_ideal else 0

    if view_is_month:
        # ── Full Month: one bar per week ──────────────────────────────────────
        bars = ""
        for wk_info in w_weeks:
            n   = inv["by_week"].get(wk_info["week_num"], 0)
            dc  = dot_color(n, cfg["weekly_min"], cfg["weekly_ideal"]) if n else "#FEE2CC"
            tc  = TX if n else "#D1D5DB"
            bars += (
                f'<div style="flex:1;text-align:center;min-width:40px">'
                f'<div style="font-size:11px;font-weight:700;color:{tc};margin-bottom:3px">'
                f'{"–" if not n else n}</div>'
                f'<div style="height:28px;background:{dc};border-radius:4px"></div>'
                f'<div style="font-size:9px;color:{TX2};margin-top:3px">'
                f'{wk_info["label"]}<br>'
                f'<span style="font-size:8px">{wk_info["detail"]}</span></div>'
                f'</div>')
        week_legend = "".join(
            f'<span style="display:flex;align-items:center;gap:3px;font-size:10px;color:{TX2}">'
            f'<span style="width:9px;height:9px;background:{lc};border-radius:2px;'
            f'display:inline-block"></span>{ll}</span>'
            for lc, ll in [
                (GRN, f'≥{cfg["weekly_ideal"]}/wk goal'),
                (ORG, f'{cfg["weekly_min"]+1}–{cfg["weekly_ideal"]-1}/wk'),
                (RED, f'≤{cfg["weekly_min"]}/wk critical'),
                ("#FEE2CC", "no cases"),
            ])
        bar_section = (
            f'<div style="font-size:9px;color:{TX2};text-transform:uppercase;'
            f'letter-spacing:.07em;margin-bottom:8px">Weekly Production</div>'
            f'<div style="display:flex;gap:8px;align-items:flex-end">{bars}</div>'
            f'<div style="margin-top:8px;display:flex;gap:7px;flex-wrap:wrap">{week_legend}</div>'
        )
    else:
        # ── Week view: one bar per work day (Mon–Fri) ─────────────────────────
        bars = ""
        for wd in w_days:
            n  = inv["by_day"].get(wd["ds"], 0)
            dc = dot_color(n, cfg["daily_min"], cfg["daily_ideal"])
            tc = TX if n else "#D1D5DB"
            bars += (
                f'<div style="flex:1;text-align:center;min-width:16px">'
                f'<div style="font-size:9px;font-weight:700;color:{tc};margin-bottom:2px">'
                f'{"–" if not n else n}</div>'
                f'<div style="height:22px;background:{dc};border-radius:3px"></div>'
                f'<div style="font-size:8px;color:{TX2};margin-top:2px">{wd["day"]}</div>'
                f'</div>')
        day_legend = "".join(
            f'<span style="display:flex;align-items:center;gap:3px;font-size:10px;color:{TX2}">'
            f'<span style="width:9px;height:9px;background:{lc};border-radius:2px;'
            f'display:inline-block"></span>{ll}</span>'
            for lc, ll in [
                (GRN, f'≥{cfg["daily_ideal"]}/day goal'),
                (ORG, f'6–{cfg["daily_ideal"]-1}/day'),
                (RED, f'≤{cfg["daily_min"]}/day critical'),
                ("#FEE2CC", "absent"),
            ])
        bar_section = (
            f'<div style="font-size:9px;color:{TX2};text-transform:uppercase;'
            f'letter-spacing:.07em;margin-bottom:6px">Daily Production</div>'
            f'<div style="display:flex;gap:4px;align-items:flex-end;overflow:hidden">{bars}</div>'
            f'<div style="margin-top:8px;display:flex;gap:7px;flex-wrap:wrap">{day_legend}</div>'
        )

    plbl = "Month total" if view_is_month else "Week total"
    prog = (
        f'<div style="display:flex;justify-content:space-between;font-size:12px;'
        f'color:{TX2};margin-bottom:3px"><span>{plbl}</span>'
        f'<span style="font-weight:700;color:{bc}">{inv["total"]} cases</span></div>'
        f'<div style="height:7px;background:{ABSC};border-radius:4px;margin-bottom:3px">'
        f'<div style="height:100%;width:{wk_pct:.1f}%;background:{bc};border-radius:4px">'
        f'</div></div>'
        f'<div style="display:flex;justify-content:space-between;font-size:10px;'
        f'color:{TX2};margin-bottom:6px">'
        f'<span>0</span><span>▲ min {w_min}</span>'
        f'<span style="color:{GRN}">ideal {w_ideal}</span></div>'
        + (f'<div style="font-size:11px;color:{TX2};margin-bottom:10px">'
           f'Month total: <b style="color:{ORG}">{inv["month_total"]}</b>'
           f' <span style="font-size:10px">({month_str})</span></div>'
           if not view_is_month else "")
    ) if not inv["support"] else (
        f'<div style="display:flex;justify-content:space-between;font-size:12px;'
        f'color:{TX2};margin-bottom:12px"><span>Cases contributed</span>'
        f'<span style="font-weight:700;color:{TX}">{inv["total"]}</span></div>'
    )

    sup_sub = (f'<div style="font-size:10px;color:{TX2}">Support role</div>'
               if inv["support"] else "")

    return f"""
    <div style="background:{CARD};border:1px solid {BORD};border-radius:14px;
                padding:14px 16px;margin-bottom:4px">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:11px">
        <div style="width:32px;height:32px;border-radius:50%;background:{OL};
                    border:2px solid {OB};display:flex;align-items:center;
                    justify-content:center;font-weight:700;color:{ORG};font-size:13px">
          {inv['name'][0]}</div>
        <div><div style="font-weight:700;font-size:14px;color:{TX}">{inv['name']}</div>
          {sup_sub}</div>
        <span style="margin-left:auto;font-size:10px;font-weight:700;background:{bb};
                     color:{bc};padding:2px 8px;border-radius:20px">{bl}</span>
      </div>
      {prog}
      {bar_section}
    </div>"""


if not invs:
    for phc in st.columns(3):
        with phc:
            st.markdown(
                f'<div style="background:{CARD};border:1px solid {BORD};'
                f'border-radius:14px;padding:16px;height:130px;display:flex;'
                f'align-items:center;justify-content:center">'
                f'<span style="color:{TX2};font-size:13px">No cases for this period</span>'
                f'</div>', unsafe_allow_html=True)
else:
    n_cols   = min(4, len(invs))
    inv_cols = st.columns(n_cols)
    for idx, inv in enumerate(invs):
        with inv_cols[idx % n_cols]:
            st.markdown(make_card(inv), unsafe_allow_html=True)
            with st.expander(f"📊 {inv['name']} — detail", expanded=False):
                ex1, ex2 = st.columns(2)
                day_vals = [inv["by_day"].get(wd["ds"], 0) for wd in w_days]
                with ex1:
                    st.markdown(
                        f'<div style="font-size:10px;font-weight:700;color:{ORG};'
                        f'text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">'
                        f'Daily</div>', unsafe_allow_html=True)
                    fig_d = go.Figure(go.Bar(
                        x=[wd["label"] for wd in w_days], y=day_vals,
                        marker_color=[dot_color(n, cfg["daily_min"], cfg["daily_ideal"])
                                      for n in day_vals],
                        marker_line_width=0))
                    fig_d.update_layout(
                        height=150, margin=dict(t=5,b=5,l=5,r=5),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        template=PLT, xaxis=dict(showgrid=False),
                        yaxis=dict(showgrid=True))
                    st.plotly_chart(fig_d, use_container_width=True,
                                    config={"displayModeBar": False},
                                    key=f"d_{idx}_{tab}_{view}")
                with ex2:
                    st.markdown(
                        f'<div style="font-size:10px;font-weight:700;color:{ORG};'
                        f'text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">'
                        f'By country</div>', unsafe_allow_html=True)
                    if not w_data.empty:
                        ic = (w_data[w_data["investigator"]==inv["name"]]
                              .groupby("country").size().sort_values().to_dict())
                        if ic:
                            fig_c = go.Figure(go.Bar(
                                x=list(ic.values()), y=list(ic.keys()),
                                orientation="h", marker_color=ORG, marker_line_width=0))
                            fig_c.update_layout(
                                height=max(120, len(ic)*28),
                                margin=dict(t=5,b=5,l=5,r=5),
                                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                template=PLT, xaxis=dict(showgrid=True),
                                yaxis=dict(showgrid=False))
                            st.plotly_chart(fig_c, use_container_width=True,
                                            config={"displayModeBar": False},
                                            key=f"c_{idx}_{tab}_{view}")
                # Month breakdown (only in week views)
                if not view_is_month and not m_data.empty:
                    im = m_data[m_data["investigator"] == inv["name"]]
                    if not im.empty:
                        st.markdown(
                            f'<div style="font-size:10px;font-weight:700;color:{ORG};'
                            f'text-transform:uppercase;letter-spacing:.08em;margin:8px 0 6px">'
                            f'Month — {month_str}</div>', unsafe_allow_html=True)
                        mbd = im.groupby("date").size()
                        md  = sorted(mbd.index); mv = [mbd[d] for d in md]
                        fig_m = go.Figure(go.Bar(
                            x=[fmt_day(d) for d in md], y=mv,
                            marker_color=[dot_color(n, cfg["daily_min"],
                                                    cfg["daily_ideal"]) for n in mv],
                            marker_line_width=0))
                        fig_m.update_layout(
                            height=150, margin=dict(t=5,b=5,l=5,r=5),
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            template=PLT, xaxis=dict(showgrid=False, tickangle=-45),
                            yaxis=dict(showgrid=True))
                        st.plotly_chart(fig_m, use_container_width=True,
                                        config={"displayModeBar": False},
                                        key=f"m_{idx}_{tab}_{view}")

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# DAILY PRODUCTION CHART
# ─────────────────────────────────────────────────────────────────────────────
with st.container(border=True):
    st.markdown(
        f'<div class="sec-lbl">{"Weekly" if view_is_month else "Daily"} Case Production — '
        f'{cfg["name"]} · {card_period}</div>',
        unsafe_allow_html=True)

    if view_is_month and w_weeks:
        # Full month: bar chart with one bar per week
        x_vals = [f"{w['label']}\n{w['detail']}" for w in w_weeks]
        y_vals = [w["total"] for w in w_weeks]
        fig_l  = go.Figure(go.Bar(
            x=x_vals, y=y_vals,
            marker_color=[GRN if n >= cfg["weekly_ideal"]
                          else (ORG if n > cfg["weekly_min"] else RED)
                          for n in y_vals],
            marker_line_width=0,
            text=y_vals, textposition="outside", textfont=dict(color=TX),
        ))
        fig_l.add_hline(y=cfg["weekly_ideal"], line_dash="dash", line_color=GRN,
                        annotation_text=f"Goal ({cfg['weekly_ideal']})",
                        annotation_position="right", annotation_font_color=GRN)
        fig_l.add_hline(y=cfg["weekly_min"], line_dash="dash", line_color=RED,
                        annotation_text=f"Critical ({cfg['weekly_min']})",
                        annotation_position="right", annotation_font_color=RED)
        fig_l.update_layout(
            height=220, margin=dict(t=10,b=10,l=10,r=90),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            template=PLT, showlegend=False,
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="rgba(249,115,22,0.1)"))
    else:
        # Week view: line chart with one point per work day
        x_vals = [wd["label"] for wd in w_days]
        y_vals = [wd["total"] for wd in w_days]
        fig_l  = go.Figure()
        if has_data and any(y_vals):
            fig_l.add_trace(go.Scatter(
                x=x_vals, y=y_vals, mode="lines+markers",
                line=dict(color=ORG, width=2.5), marker=dict(color=ORG, size=8),
                fill="tozeroy", fillcolor="rgba(249,115,22,0.12)"))
        else:
            fig_l.add_trace(go.Scatter(
                x=x_vals, y=[0]*len(x_vals), mode="lines",
                line=dict(color=BORD, width=2)))
        fig_l.add_hline(y=cfg["daily_ideal"], line_dash="dash", line_color=GRN,
                        annotation_text=f"Goal ({cfg['daily_ideal']})",
                        annotation_position="right", annotation_font_color=GRN)
        fig_l.add_hline(y=cfg["daily_min"], line_dash="dash", line_color=RED,
                        annotation_text=f"Critical ({cfg['daily_min']})",
                        annotation_position="right", annotation_font_color=RED)
        fig_l.update_layout(
            height=220, margin=dict(t=10,b=10,l=10,r=90),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            template=PLT, showlegend=False,
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="rgba(249,115,22,0.1)"))

    st.plotly_chart(fig_l, use_container_width=True, config={"displayModeBar": False})

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# BOTTOM CHARTS
# ─────────────────────────────────────────────────────────────────────────────
def _no_data():
    return (f'<p style="color:{TX2};font-size:13px;padding:20px 0">'
            f'No cases for this period</p>')

bc1, bc2 = st.columns(2)
with bc1:
    with st.container(border=True):
        st.markdown('<div class="sec-lbl">Cases by Country</div>',
                    unsafe_allow_html=True)
        if by_country:
            fig_c = go.Figure(go.Bar(
                x=list(by_country.values()), y=list(by_country.keys()),
                orientation="h", marker_color=ORG, marker_line_width=0,
                text=list(by_country.values()), textposition="outside",
                textfont=dict(color=TX)))
            fig_c.update_layout(
                height=max(220, len(by_country)*30),
                margin=dict(t=5,b=5,l=10,r=40),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                template=PLT,
                xaxis=dict(showgrid=True, gridcolor="rgba(249,115,22,0.1)"),
                yaxis=dict(showgrid=False, autorange="reversed"))
            st.plotly_chart(fig_c, use_container_width=True,
                            config={"displayModeBar": False})
        else:
            st.markdown(_no_data(), unsafe_allow_html=True)

with bc2:
    with st.container(border=True):
        st.markdown('<div class="sec-lbl">Cases by Investigator</div>',
                    unsafe_allow_html=True)
        if by_inv_stat:
            fig_i = go.Figure(go.Bar(
                x=[i["total"] for i in by_inv_stat],
                y=[i["name"]  for i in by_inv_stat],
                orientation="h", marker_color=ORG, marker_line_width=0,
                text=[f"{i['total']} ({i['pct']}%)" for i in by_inv_stat],
                textposition="outside", textfont=dict(color=TX)))
            fig_i.update_layout(
                height=max(220, len(by_inv_stat)*40),
                margin=dict(t=5,b=5,l=10,r=90),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                template=PLT,
                xaxis=dict(showgrid=True, gridcolor="rgba(249,115,22,0.1)"),
                yaxis=dict(showgrid=False, autorange="reversed"))
            st.plotly_chart(fig_i, use_container_width=True,
                            config={"displayModeBar": False})
        else:
            st.markdown(_no_data(), unsafe_allow_html=True)
